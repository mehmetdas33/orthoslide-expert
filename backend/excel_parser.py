"""
Excel Parser for Cephalometric Data
OrthoSlide Expert V2

Reads .xlsx file: Column B, starting at Row 3
Maps row indices to named measurement keys.
"""
import os
import pandas as pd
from ceph_logic import EXCEL_ROW_MAP


def _ensure_xlsx(file_path: str) -> str:
    """
    Accepts .xls or .xlsx files and always returns a path to a valid .xlsx.

    Handles four formats:
      1. Already .xlsx                → return as-is
      2. ZIP-based (xlsx renamed .xls) → copy with .xlsx extension
      3. HTML-based XLS (some tools export measurement tables as HTML with .xls)
         → parse with pd.read_html, rebuild proper xlsx layout
      4. Binary XLS (BIFF format)    → convert cell-by-cell with xlrd
      Fallback: pandas xlrd engine
    """
    if file_path.lower().endswith('.xlsx'):
        return file_path

    xlsx_path = os.path.splitext(file_path)[0] + '_converted.xlsx'

    # ── Case 1: file is a ZIP (xlsx content with wrong extension) ──
    import zipfile
    if zipfile.is_zipfile(file_path):
        import shutil
        shutil.copy2(file_path, xlsx_path)
        return xlsx_path

    # ── Case 2: HTML-based XLS ──
    # Detect by checking first bytes for BOM + '<' or just '<'
    try:
        with open(file_path, 'rb') as _f:
            _head = _f.read(64)
        _is_html = _head.lstrip(b'\xef\xbb\xbf').lstrip()[:1] == b'<'
    except Exception:
        _is_html = False

    if _is_html:
        try:
            import openpyxl

            tables = pd.read_html(file_path, encoding='utf-8', flavor='lxml')
            wb = openpyxl.Workbook()
            ws = wb.active

            # Table 0 = patient info row: col0=name, col4=gender
            if tables:
                t0 = tables[0]
                if len(t0) > 0:
                    val_name = t0.iloc[0, 0]
                    ws.cell(1, 1, str(val_name) if pd.notna(val_name) else '')
                    if t0.shape[1] >= 5:
                        val_gender = t0.iloc[0, 4]
                        if pd.notna(val_gender):
                            ws.cell(1, 5, str(val_gender))

            # Tables 2+ = one measurement per table (table 1 is the column header)
            data_tables = [t for t in tables[2:] if len(t) > 0]
            for offset, tbl in enumerate(data_tables):
                excel_row = 3 + offset           # Row 3 = parse_excel start
                label = tbl.iloc[0, 0]
                ws.cell(excel_row, 1, str(label) if pd.notna(label) else '')
                if tbl.shape[1] > 1:
                    raw = tbl.iloc[0, 1]
                    if pd.notna(raw):
                        try:
                            ws.cell(excel_row, 2, float(raw))
                        except (ValueError, TypeError):
                            ws.cell(excel_row, 2, str(raw))

            wb.save(xlsx_path)
            return xlsx_path
        except Exception as e:
            raise ValueError(f"HTML-XLS dönüştürme hatası: {e}")

    # ── Case 3: Binary XLS (BIFF) via xlrd ──
    try:
        import xlrd
        import openpyxl

        wb_in = xlrd.open_workbook(file_path)
        ws_in = wb_in.sheet_by_index(0)
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active

        for row in range(ws_in.nrows):
            for col in range(ws_in.ncols):
                ws_out.cell(row + 1, col + 1, ws_in.cell(row, col).value)

        wb_out.save(xlsx_path)
        return xlsx_path
    except ImportError:
        pass  # xlrd not installed — fall through
    except Exception as e:
        raise ValueError(f"XLS dönüştürme hatası: {e}. Dosyayı .xlsx olarak kaydedin.")

    # ── Fallback: pandas xlrd engine ──
    try:
        df = pd.read_excel(file_path, header=None, engine='xlrd')
        df.to_excel(xlsx_path, index=False, header=False, engine='openpyxl')
        return xlsx_path
    except Exception as e:
        raise ValueError(
            f"XLS dosyası okunamadı. Excel'de açıp 'Farklı Kaydet → .xlsx' ile kaydedin. ({e})"
        )


def parse_excel(file_path: str) -> dict:
    """
    Parse an Excel file and return a dict of measurement_key → value.

    Expected layout:
      - Column A: measurement labels (optional, for reference)
      - Column B: measurement values
      - Data starts at Row 3 (0-indexed row 2 in pandas)
    """
    file_path = _ensure_xlsx(file_path)
    df = pd.read_excel(file_path, header=None, engine="openpyxl")

    data = {}
    start_row = 2  # Row 3 in Excel (0-indexed)

    for i, key in enumerate(EXCEL_ROW_MAP):
        row_idx = start_row + i
        if row_idx < len(df):
            raw_value = df.iloc[row_idx, 1]  # Column B (index 1)
            if pd.notna(raw_value):
                try:
                    data[key] = float(raw_value)
                except (ValueError, TypeError):
                    data[key] = str(raw_value)
            else:
                data[key] = None
        else:
            data[key] = None

    return data


import re

def parse_patient_info(file_path: str) -> dict:
    """
    Parse patient info from the Excel file.
    Row 1, Col A (A1): Patient name in 'Surname, Name' format.
    Row 1, Col E (E1): Gender ('Female' or 'Male').
    """
    try:
        file_path = _ensure_xlsx(file_path)
        df = pd.read_excel(file_path, header=None, engine="openpyxl", nrows=2, usecols="A:E")
        info = {}
        
        # A1 is index [0, 0]
        if len(df) > 0 and pd.notna(df.iloc[0, 0]):
            raw_name = str(df.iloc[0, 0]).strip()
            if "," in raw_name:
                parts = raw_name.split(",", 1)
                name_str = f"{parts[1].strip()} {parts[0].strip()}"
            else:
                name_str = raw_name
            # Remove anything inside ()
            name_str = re.sub(r'\(.*?\)', '', name_str).strip()
            # Clean up extra spaces
            name_str = re.sub(r'\s+', ' ', name_str)
            info["patient_name"] = name_str
        else:
            info["patient_name"] = ""
            
        # E1 is index [0, 4]
        if df.shape[1] >= 5 and pd.notna(df.iloc[0, 4]):
            info["gender"] = str(df.iloc[0, 4]).strip()
        else:
            info["gender"] = ""
            
        return info
    except Exception as e:
        print(f"Error parsing patient info: {e}")
        return {"patient_name": "", "gender": ""}
